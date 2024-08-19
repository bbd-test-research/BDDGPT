from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup


def extract_text_from_url(url):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()  # Levanta um erro para códigos de status HTTP 4xx/5xx
        soup = BeautifulSoup(response.content, 'html.parser')
        text = soup.get_text(separator=' ', strip=True)  # Extraindo e limpando o texto
        return text
    except requests.exceptions.RequestException as e:
        print(f'Erro na solicitação: {e}')
        return None


def copiar_coluna_j_para_f_iterativo(arquivo_excel):
    workbook = load_workbook(filename=arquivo_excel)
    sheet = workbook.active
    for row in range(2, sheet.max_row + 1):
        valor_coluna_f = sheet[f'F{row}'].value
        if valor_coluna_f is not None:
            continue  # Se já existe valor na coluna F, pula para a próxima linha
        valor_coluna_j = sheet[f'J{row}'].value
        valor_coluna_f = processar_linha(valor_coluna_j)
        sheet[f'F{row}'] = valor_coluna_f
    workbook.save(filename=arquivo_excel)


def processar_linha(url):
    return "primeira letra:" + extract_text_from_url(url)[0]
